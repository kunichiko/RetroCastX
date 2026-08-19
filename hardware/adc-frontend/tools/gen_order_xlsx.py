#!/usr/bin/env python3
"""調達管理用の Excel を生成する。どの部品をどこで買ったかを追跡する。

## 何が出るか

    orders/<tag>/purchase-tracker.xlsx

  シート「調達管理」  部品を縦に並べ、ステータスをドロップダウンで選べる表
  シート「発注先まとめ」 仕入先ごとの品目数・確認済み金額
  シート「注意事項」  発注前に確認すべきこと(order_sourcing.json の note)

## ステータス

ドロップダウンで選ぶ。色が自動で付く(条件付き書式):

    未発注        赤
    発注済み      黄
    到着済み      緑
    手持ち在庫あり 緑
    不要          灰

## 既存の入力を保つ

★再生成しても**状態・発注先・注文番号・備考は引き継がれる**。既存ファイルがあれば
デジグネータ列をキーに読み戻してから書き直す。部品構成が変わっても
入力済みのチェックが消えない。

発注先は CSV のファイル名から機械的に決めるが、**利用者が手で変えたらそちらを優先する**
(例: DigiKey 在庫切れで AliExpress に切り替えた場合)。上書きした件は実行時に報告する。
★2026-08-19、この引き継ぎが無かったため U9 の発注先(DigiKey→AliExpress の手編集)を
  再生成で消してしまった。

## 使い方

    python3 tools/gen_order_xlsx.py --tag v0.9.0 --boards 2

元データ:
    orders/<tag>/digikey-bom.csv, lcsc-bom.csv   発注数と品番
    tools/order_sourcing.json                     仕入先・注意事項
    tools/prices.json                             確認済みの単価
"""
import argparse
import csv
import json
import pathlib
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = pathlib.Path(__file__).resolve().parent.parent
SOURCING = ROOT / "tools/order_sourcing.json"
PRICES = ROOT / "tools/prices.json"

STATUSES = ["未発注", "発注済み", "到着済み", "手持ち在庫あり", "不要"]
STATUS_FILL = {
    "未発注": "FFC7CE",          # 赤
    "発注済み": "FFEB9C",        # 黄
    "到着済み": "C6EFCE",        # 緑
    "手持ち在庫あり": "C6EFCE",  # 緑
    "不要": "D9D9D9",            # 灰
}
# 仕入先ごとの色
VENDOR_FILL = {
    "DigiKey": "DDEBF7",
    "LCSC": "FFF2CC",
    "AliExpress": "FCE4D6",
    "秋月電子等": "E2EFDA",
}

HEAD = "1F4E79"      # 見出しの背景
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = [
    ("状態", 15),
    ("デジグネータ", 26),
    ("品名/内容", 34),
    ("発注先", 12),
    ("品番", 30),
    ("メーカー", 20),
    ("発注数", 8),
    ("単価", 12),
    ("小計(円)", 11),
    ("注文番号/日付", 18),
    ("備考", 40),
]


# ★発注先は**どの発注CSVに載っているか**で決める。
#   order_sourcing.json の status から引こうとすると、status=ok の品目は alt_mpn を
#   持たないため MPN で照合できず、全部が「秋月電子等」に落ちる(2026-08-19に踏んだ)。
#   「digikey-bom.csv に載っている = DigiKey で買う」が定義そのものなので、それに従う。
CSV_VENDOR = {"digikey-bom.csv": "DigiKey", "lcsc-bom.csv": "LCSC"}
# CSV 上は LCSC 枠だが実際には別の店で買うもの
VENDOR_OVERRIDE = {
    "C54408": "AliExpress",   # LCSC 在庫切れ。手半田分は AliExpress
}


def vendor_of(csv_name, key, refs):
    if key in VENDOR_OVERRIDE:
        return VENDOR_OVERRIDE[key]
    if refs.startswith("PG") or "module" in refs:
        return "AliExpress"   # ポゴピン / i5モジュール
    return CSV_VENDOR.get(csv_name, "秋月電子等")


def load_rows(tag, boards):
    src = json.loads(SOURCING.read_text())
    pj = json.loads(PRICES.read_text())
    prices, fx = pj["prices"], pj["fx"]
    parts = src["parts"]
    rows = []

    d = ROOT / "orders" / tag
    for fn, keycol, refcol in (("digikey-bom.csv", "Manufacturer Part Number", "Customer Reference"),
                               ("lcsc-bom.csv", "LCSC Part Number", "Designator")):
        f = d / fn
        if not f.exists():
            continue
        for x in csv.DictReader(f.open()):
            key = x[keycol].strip()
            mpn = x.get("Manufacturer Part Number", "").strip()
            # 注意事項(note)を引くためだけに sourcing を探す。発注先には使わない
            info = parts.get(key)
            if info is None and mpn:
                info = next((v for v in parts.values() if v.get("alt_mpn") == mpn), None)
            vendor = vendor_of(fn, key, x[refcol])
            p = prices.get(key) or prices.get(mpn)
            unit = p["unit"] if p else None
            yen = fx["USD_JPY_digikey"] if vendor == "DigiKey" else fx["USD_JPY_lcsc"]
            qty = int(x["Quantity"])
            rows.append({
                "refs": x[refcol],
                "desc": x.get("Description", ""),
                "vendor": vendor,
                "mpn": mpn or key,
                "mfr": x.get("Manufacturer", ""),
                "qty": qty,
                "unit_usd": unit,
                "sub_jpy": unit * qty * yen if unit else None,
                "note": (info or {}).get("note", ""),
            })

    # extra のうち CSV に出ていないもの(汎用品)を足す
    seen = {r["refs"] for r in rows}
    for e in src["extra"]:
        if e["ref"] in seen:
            continue
        rows.append({
            "refs": e["ref"], "desc": e["desc"][:200],
            "vendor": vendor_of("", "", e["ref"]),
            "mpn": e["mpn"], "mfr": e.get("mfr", ""),
            "qty": e["qty"] * boards, "unit_usd": None, "sub_jpy": None,
            "note": "",
        })
    order = {"DigiKey": 0, "LCSC": 1, "AliExpress": 2, "秋月電子等": 3}
    rows.sort(key=lambda r: (order.get(r["vendor"], 9), -(r["sub_jpy"] or 0), r["refs"]))
    return rows, fx


def read_existing(path):
    """既存ファイルから状態・注文番号・備考を読み戻す(デジグネータをキーに)。"""
    if not path.exists():
        return {}
    try:
        wb = load_workbook(path)
        ws = wb["調達管理"]
    except Exception as e:
        print(f"  既存ファイルを読めなかった({e})。新規作成する", file=sys.stderr)
        return {}
    names = [c.value for c in ws[2]]
    idx = {n: i for i, n in enumerate(names) if n}
    keep = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        ref = row[idx["デジグネータ"]] if "デジグネータ" in idx else None
        if not ref:
            continue
        keep[str(ref)] = {
            "状態": row[idx["状態"]] if "状態" in idx else None,
            # ★発注先も引き継ぐ。CSV から機械的に決めた値を利用者が手で変えることが
            #   あるため(2026-08-19: U9 を DigiKey → AliExpress に変えたのを
            #   再生成で消してしまった)。手編集を優先し、上書きした件数を報告する。
            "発注先": row[idx["発注先"]] if "発注先" in idx else None,
            "注文番号/日付": row[idx.get("注文番号/日付", -1)] if "注文番号/日付" in idx else None,
            "備考": row[idx["備考"]] if "備考" in idx else None,
        }
    return keep


def style_header(ws, title, subtitle):
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["B1"] = subtitle
    ws["B1"].font = Font(size=9, color="808080")
    for i, (name, w) in enumerate(COLS, start=1):
        c = ws.cell(row=2, column=i, value=name)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEAD)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BOX
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A3"


def build(tag, boards):
    out = ROOT / "orders" / tag / "purchase-tracker.xlsx"
    keep = read_existing(out)
    rows, fx = load_rows(tag, boards)

    wb = Workbook()
    ws = wb.active
    ws.title = "調達管理"
    style_header(ws, f"RetroCast X 調達管理 ({tag} / {boards}枚分)",
                 "tools/gen_order_xlsx.py が生成。状態・注文番号・備考は再生成しても引き継がれます")

    overridden = []
    for i, r in enumerate(rows):
        row = 3 + i
        prev = keep.get(r["refs"], {})
        # 手で変えた発注先があればそれを使う(単価の換算レートもそちらに合わせる)
        pv = prev.get("発注先")
        if pv and pv in VENDOR_FILL and pv != r["vendor"]:
            overridden.append((r["refs"], r["vendor"], pv))
            r["vendor"] = pv
        vals = [
            prev.get("状態") or "未発注",
            r["refs"], r["desc"][:120], r["vendor"], r["mpn"], r["mfr"], r["qty"],
            f'${r["unit_usd"]:.4f}' if r["unit_usd"] else "未確認",
            round(r["sub_jpy"]) if r["sub_jpy"] else None,
            prev.get("注文番号/日付") or "",
            prev.get("備考") or "",
        ]
        for j, v in enumerate(vals, start=1):
            c = ws.cell(row=row, column=j, value=v)
            c.border = BOX
            c.alignment = Alignment(vertical="top",
                                    wrap_text=(j in (3, 11)),
                                    horizontal="center" if j in (1, 4, 7) else None)
        ws.cell(row=row, column=4).fill = PatternFill(
            "solid", fgColor=VENDOR_FILL.get(r["vendor"], "FFFFFF"))
        if r["sub_jpy"]:
            ws.cell(row=row, column=9).number_format = "#,##0"
        else:
            ws.cell(row=row, column=8).font = Font(color="C00000")
        ws.row_dimensions[row].height = 30

    last = 2 + len(rows)

    # 状態のドロップダウン
    dv = DataValidation(type="list", formula1='"' + ",".join(STATUSES) + '"',
                        allow_blank=False, showDropDown=False)
    dv.error = "一覧から選んでください"
    dv.errorTitle = "無効な値"
    ws.add_data_validation(dv)
    dv.add(f"A3:A{last}")

    # 状態に応じて行全体を色付け
    for st, color in STATUS_FILL.items():
        ws.conditional_formatting.add(
            f"A3:K{last}",
            FormulaRule(formula=[f'$A3="{st}"'],
                        fill=PatternFill("solid", fgColor=color), stopIfTrue=False))

    # 合計行
    t = last + 2
    ws.cell(row=t, column=1, value="合計").font = Font(bold=True)
    ws.cell(row=t, column=7, value=f"=SUM(G3:G{last})").font = Font(bold=True)
    ws.cell(row=t, column=9, value=f"=SUM(I3:I{last})").font = Font(bold=True)
    ws.cell(row=t, column=9).number_format = "#,##0"
    ws.cell(row=t, column=10, value="★単価「未確認」の品目は小計に入っていません")
    ws.cell(row=t, column=10).font = Font(color="C00000", size=9)

    t2 = t + 2
    ws.cell(row=t2, column=1, value="進捗").font = Font(bold=True)
    for k, st in enumerate(STATUSES):
        ws.cell(row=t2 + k, column=2, value=st)
        ws.cell(row=t2 + k, column=3,
                value=f'=COUNTIF($A$3:$A${last},"{st}") & " / {len(rows)} 品目"')

    # --- 発注先まとめ ---
    ws2 = wb.create_sheet("発注先まとめ")
    ws2["A1"] = "発注先ごとの内訳"
    ws2["A1"].font = Font(bold=True, size=14)
    for i, h in enumerate(["発注先", "品目数", "総個数", "確認済み金額(円)", "アップロード用ファイル"], start=1):
        c = ws2.cell(row=2, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEAD)
        c.border = BOX
    for w, col in zip((16, 10, 10, 18, 30), "ABCDE"):
        ws2.column_dimensions[col].width = w
    files = {"DigiKey": "orders/<tag>/digikey-bom.csv",
             "LCSC": "orders/<tag>/lcsc-bom.csv",
             "AliExpress": "(手動。RJ45・ポゴピン・i5モジュール)",
             "秋月電子等": "(手動。ピンヘッダ・ネジ・PLR135/T)"}
    r0 = 3
    for v in ("DigiKey", "LCSC", "AliExpress", "秋月電子等"):
        g = [r for r in rows if r["vendor"] == v]
        if not g:
            continue
        known = sum(r["sub_jpy"] for r in g if r["sub_jpy"])
        for i, val in enumerate([v, len(g), sum(r["qty"] for r in g),
                                 round(known) if known else "未確認",
                                 files.get(v, "")], start=1):
            c = ws2.cell(row=r0, column=i, value=val)
            c.border = BOX
        ws2.cell(row=r0, column=1).fill = PatternFill("solid", fgColor=VENDOR_FILL.get(v, "FFFFFF"))
        ws2.cell(row=r0, column=4).number_format = "#,##0"
        r0 += 1
    ws2.cell(row=r0 + 1, column=1,
             value=f'為替: DigiKey は USD x {fx["USD_JPY_digikey"]}、LCSC/AliExpress は USD x {fx["USD_JPY_lcsc"]}。'
                   "確定額は各社の BOM ツールで出すこと。").font = Font(size=9, color="808080")

    # --- 注意事項 ---
    ws3 = wb.create_sheet("注意事項")
    ws3["A1"] = "★発注前に確認すること"
    ws3["A1"].font = Font(bold=True, size=14)
    for i, h in enumerate(["部品", "品番", "注意"], start=1):
        c = ws3.cell(row=2, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=HEAD)
        c.border = BOX
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 30
    ws3.column_dimensions["C"].width = 110
    r0 = 3
    for r in rows:
        if not r["note"]:
            continue
        for i, v in enumerate([r["refs"].split(",")[0], r["mpn"], r["note"]], start=1):
            c = ws3.cell(row=r0, column=i, value=v)
            c.border = BOX
            c.alignment = Alignment(vertical="top", wrap_text=(i == 3))
        ws3.row_dimensions[r0].height = 90
        r0 += 1
    ws3.freeze_panes = "A3"

    wb.save(out)
    return out, rows, keep, overridden


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--tag", default="v0.9.0")
    p.add_argument("--boards", type=int, default=2)
    a = p.parse_args()
    out, rows, keep, overridden = build(a.tag, a.boards)
    print(f"{out.relative_to(ROOT)} に出力しました ({a.boards}枚分)")
    print(f"  {len(rows)}品目")
    for v in ("DigiKey", "LCSC", "AliExpress", "秋月電子等"):
        g = [r for r in rows if r["vendor"] == v]
        if g:
            print(f"    {v:<12} {len(g):>3}品目 {sum(r['qty'] for r in g):>5}個")
    if keep:
        print(f"  既存の入力を {len(keep)}行から引き継いだ(状態/発注先/注文番号/備考)")
    for refs, auto, manual in overridden:
        print(f"    発注先は手編集を優先: {refs} {auto} → {manual}")
    unknown = sum(1 for r in rows if not r["unit_usd"])
    if unknown:
        print(f"  ★単価未確認 {unknown}品目(小計に入っていない)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
